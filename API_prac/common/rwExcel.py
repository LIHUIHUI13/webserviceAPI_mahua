# !/usr/bin/python3
# coding:utf-8 
# Author : mahua
# Email : lihh3721@gmail.com
# Time : 2019/4/15 12:34 AM
# FileName : rwExcel.py

from openpyxl import load_workbook
from API_prac.common import http_request

#测试用例类
class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None

class RWExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
    def readExcel(self):#读取excel内容，每一行的用字典存储，最后所有的行用例再添加到一个列表里
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        max_row = sheet.max_row
        cases=[]
        for r in range(2,max_row+1):
            case=Case() #生成一个测试类的实例
            case.case_id = sheet.cell(row=r,column=1).value
            case.title = sheet.cell(row=r,column=2).value
            case.url = sheet.cell(row=r,column=3).value
            case.data = sheet.cell(row=r,column=4).value
            case.method = sheet.cell(row=r,column=5).value
            case.expected = sheet.cell(row=r,column=6).value
            case.sql = sheet.cell(row=r,column=9).value
            cases.append(case)
        wb.close()
        return cases
    def writeExcel(self,row,actual,result):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row,7).value = actual
        sheet.cell(row,8).value = result
        wb.save(filename=self.file_name)
        wb.close()


if __name__ == '__main__':
    pass
    # from API_3.common import contants
    #
    # sheet_name = ['register','login','recharge']
    # request = http_request.HttpRequest_session()  # 创建一个HttpRequest_session类的实例
    # for sheetName in sheet_name:
    #     rExcel = RWExcel(contants.case_file,sheetName) #从common/contants获取case_file
    #     cases = rExcel.readExcel()
    #
    #     for case in cases:
    #         print(case.__dict__) #返回所有属性的字典
    #         resp = request.http_request(case.method,case.url,case.data)#调用http_request方法
    #         actual = resp.text #获取actual的值
    #         if actual == case.expected:#注册成功和充值成功的还不知道怎么解决
    #             rExcel.writeExcel(case.case_id+1,actual,'PASS')
    #         else:
    #             rExcel.writeExcel(case.case_id+1,actual,'FAIL')
    # request.close()#关闭session


